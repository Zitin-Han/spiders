import os
import time
import pandas as pd
import pyautogui
import pyperclip
import xlwings as xw
import logging
import sys
import psutil
import ctypes
import warnings
from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# 加载 .env 环境变量
load_dotenv()

warnings.filterwarnings('ignore')

# ================= 1. 日志配置中心 =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("task_log.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ================= 配置中心（从 .env 读取，避免硬编码敏感信息）=================
USERNAME = os.getenv("LINGDONG_USERNAME")
PASSWORD = os.getenv("LINGDONG_PASSWORD")
BASE_URL = "https://bll.lingdongsz.com/uranus/#/login"
DOWNLOAD_DIR = os.getenv("DOWNLOAD_DIR", r"C:\Users\kangjia\Downloads")
LOCAL_FILE = os.getenv("LOCAL_FILE", r"D:\study\work\蓝鲸BI项目\店铺每日数据更新.xlsx")
SHEET_NAME = "数据更新表"
KDOCS_URL = os.getenv("KDOCS_URL", "https://www.kdocs.cn/")
SEARCH_NAME = os.getenv("KDOCS_SEARCH_NAME", "店铺每日数据更新")
TARGET_FILE = "店铺每日数据更新.xlsx"
pyautogui.FAILSAFE = True

# ================= 工具函数 =================
def force_close_excel(target_file_name):
    logger.info(f"正在检查文件是否被占用: {target_file_name}")
    found = False
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if proc.info['name'] and 'EXCEL' in proc.info['name'].upper():
                for fetched_file in proc.open_files():
                    if target_file_name in fetched_file.path:
                        logger.warning(f"- 检测到文件被进程 {proc.info['pid']} 占用，正在强制关闭...")
                        proc.kill()
                        found = True
                        time.sleep(1)
                        break
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue
    if not found:
        logger.info("- 文件未被占用，环境清洁")
    else:
        logger.info("- 占用进程已清理完毕")

def select_time_range(driver, wait, text):
    logger.info(f"选择时间范围：{text}")
    select_input = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//form[contains(@class,'select_time')]//input[@placeholder='请选择']"
    )))
    driver.execute_script("arguments[0].click();", select_input)
    option = wait.until(EC.element_to_be_clickable((
        By.XPATH, f"//div[contains(@class,'el-select-dropdown')]//span[text()='{text}']"
    )))
    driver.execute_script("arguments[0].click();", option)
    time.sleep(2)

def get_latest_file(folder, keyword):
    files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith((".xlsx", ".xls")) and keyword in f]
    if not files: raise Exception(f"未找到包含【{keyword}】的文件")
    return max(files, key=os.path.getmtime)

def get_latest_ads_files(folder):
    files = sorted(
        [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith((".xlsx", ".xls")) and "广告流量" in f],
        key=os.path.getmtime, reverse=True)
    if len(files) < 2: raise Exception("广告流量文件少于2个")
    return files[0], files[1]

def write_dataframe(ws, df, columns, start_col, start_row=2):
    # 提取需要写入的数据
    data = df[columns]
    num_rows = len(data)
    num_cols = len(columns)
    
    logger.info(f"正在写入列(起始位置:{start_col})，新数据行数：{num_rows}")

    if ws.max_row >= start_row:
        for r in range(start_row, ws.max_row + 1):
            for j in range(num_cols):
                ws.cell(row=r, column=start_col + j).value = None

    for i, row in data.iterrows():
        for j, col in enumerate(columns):
            # 将处理后的数值写入单元格
            ws.cell(row=start_row + i, column=start_col + j).value = row[col]
            
    logger.info(f"列(起始位置:{start_col}) 替换完成。")

def copy_local_excel_to_clipboard():
    logger.info("正在读取本地Excel并保留格式复制...")
    df = pd.read_excel(LOCAL_FILE, sheet_name=SHEET_NAME)
    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(LOCAL_FILE)
        sht = wb.sheets[SHEET_NAME]
        sht.used_range.copy()
        logger.info("数据及格式已成功复制到系统剪贴板！")
        return df, app, wb
    except Exception as e:
        app.quit()
        raise e

def clear_clipboard():
    try:
        ctypes.windll.user32.OpenClipboard(None)
        ctypes.windll.user32.EmptyClipboard()
        ctypes.windll.user32.CloseClipboard()
        logger.info("- 系统剪贴板已强制清空，预防弹窗")
    except Exception as e:
        logger.warning(f"- 清空剪贴板失败: {e}")